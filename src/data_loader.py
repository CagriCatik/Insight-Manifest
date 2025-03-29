# src/data_loader.py
import os
import logging
import pandas as pd
import openpyxl
import yaml

logger = logging.getLogger(__name__)

def load_config(config_path="config/config.yaml"):
    """Load configuration from YAML, converting relative paths to absolute paths based on the config file location."""
    with open(config_path, "r") as f:
        config = yaml.safe_load(f)
    
    # Determine the base directory of the config file
    base_dir = os.path.dirname(os.path.abspath(config_path))
    
    # List of keys that represent file paths
    path_keys = ["excel_file", "heatmap_output", "latex_template", "report_output_dir"]
    
    for key in path_keys:
        if key in config and not os.path.isabs(config[key]):
            config[key] = os.path.abspath(os.path.join(base_dir, config[key]))
    
    logger.debug("Configuration loaded: %s", config)
    return config

def read_excel_file(file_path, sheet_name):
    """Read and validate the Excel file, returning a DataFrame with test coverage data."""
    if not os.path.exists(file_path):
        logger.error("Excel file does not exist: %s", file_path)
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path)
        logger.info("Available sheets: %s", wb.sheetnames)
        if sheet_name not in wb.sheetnames:
            logger.error("Sheet '%s' not found in workbook.", sheet_name)
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    except Exception as e:
        logger.exception("Error loading workbook: %s", e)
        raise

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)
        logger.info("Excel file read successfully with shape: %s", df.shape)
    except Exception as e:
        logger.exception("Error reading Excel file into DataFrame: %s", e)
        raise
    return df
